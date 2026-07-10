"""
import_inventario.py
-------------------
Script standalone para importar dados históricos da planilha
"Inventário Continuo2.xlsx" para a tabela inventario_contagens do Supabase.

Dependências: openpyxl, requests (já instalados)
"""

import sys
import json
import datetime
import requests
import openpyxl

# ---------------------------------------------------------------------------
# Configuração Supabase
# ---------------------------------------------------------------------------
SUPABASE_URL = "https://lgpsqwgvepdfilknggec.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImxncHNxd2d2ZXBkZmlsa25nZ2VjIiwicm9sZSI6"
    "InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MDM3ODY0NywiZXhwIjoyMDk1OTU0NjQ3fQ"
    ".9gnrlLZshPQq5effMXO-Km1z_HWvWonme9nWaUjm09A"
)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

# Caminho da planilha
XLSX_PATH = r"C:\Users\tassio.santana\Downloads\Inventário Continuo2.xlsx"

# Tamanho do lote para inserção
BATCH_SIZE = 50

# ---------------------------------------------------------------------------
# Mapa de normalização de operadores
# ---------------------------------------------------------------------------
OPERADOR_MAP = {
    # Variações de Mirailton
    "mirailton": "Mirailton",
    "Mirailton": "Mirailton",
    "MIRAILTON": "Mirailton",
    "miailton": "Mirailton",
    "mirailtom": "Mirailton",
    "miraiton": "Mirailton",
    "mitailton": "Mirailton",
    "mitrailton": "Mirailton",
    # Variações de Italo
    "italo": "Italo",
    "Italo": "Italo",
    "ITALO": "Italo",
}

# ---------------------------------------------------------------------------
# Funções auxiliares
# ---------------------------------------------------------------------------

def normalizar_operador(val):
    """Normaliza o nome do operador usando o mapa de variações."""
    if val is None:
        return None
    raw = str(val).strip()
    return OPERADOR_MAP.get(raw, raw)  # fallback: retorna o valor original


def normalizar_data(val, row_num):
    """
    Converte o valor da célula data para string ISO (YYYY-MM-DD).
    Aceita datetime, date ou string com formatos brasileiros.
    Retorna (str_iso, warning) ou (None, motivo_erro).
    """
    if val is None:
        return None, f"Linha {row_num}: data vazia"

    # Já é datetime/date do openpyxl
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d"), None

    # Tenta tratar como string
    raw = str(val).strip().rstrip("'\"")  # remove apóstrofo final (ex: "22/05/2026'")

    # Tenta formatos BR: DD/MM/YYYY e variações
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            dt = datetime.datetime.strptime(raw, fmt)
            return dt.strftime("%Y-%m-%d"), None
        except ValueError:
            pass

    # Tenta corrigir typos comuns como "21/052/2026" → "21/05/2026"
    import re
    match = re.match(r"(\d{1,2})/0?(\d{1,2})/(\d{4})", raw)
    if match:
        d, m, y = match.groups()
        try:
            dt = datetime.datetime(int(y), int(m), int(d))
            print(f"  [WARN] Linha {row_num}: data corrigida de '{val}' → '{dt.strftime('%Y-%m-%d')}'")
            return dt.strftime("%Y-%m-%d"), None
        except ValueError:
            pass

    return None, f"Linha {row_num}: data inválida '{val}'"


def normalizar_codigo(val):
    """Converte int/float/str para string limpa e maiúscula."""
    if val is None:
        return None
    if isinstance(val, float):
        val = int(val)
    return str(val).strip().upper()


def normalizar_lote(val):
    """Converte int/float/str para string limpa e maiúscula."""
    if val is None:
        return None
    if isinstance(val, float):
        val = int(val)
    return str(val).strip().upper()


def to_int(val, default=None):
    """Converte val para int, retorna default se None ou inválido."""
    if val is None:
        return default
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def supabase_post(endpoint, payload):
    """Faz POST para o Supabase REST API e retorna (ok, dados/erro)."""
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    resp = requests.post(url, headers=HEADERS, json=payload, timeout=30)
    if resp.status_code in (200, 201):
        return True, resp.json()
    return False, f"HTTP {resp.status_code}: {resp.text}"


# ---------------------------------------------------------------------------
# 1. Criar ciclo histórico
# ---------------------------------------------------------------------------

def criar_ciclo():
    """Cria o ciclo histórico no Supabase e retorna o ID."""
    print("=" * 60)
    print("PASSO 1: Criando ciclo histórico no Supabase...")

    ciclo = {
        "nome": "Histórico Importado — Maio/Jun 2026",
        "data_abertura": "2026-05-19",
        "data_fechamento": "2026-06-12",
        "status": "ENCERRADO",
        "meta_itens": None,
    }

    ok, resultado = supabase_post("inventario_ciclos", ciclo)
    if not ok:
        print(f"  [ERRO] Falha ao criar ciclo: {resultado}")
        sys.exit(1)

    if isinstance(resultado, list):
        ciclo_id = resultado[0]["id"]
    else:
        ciclo_id = resultado["id"]

    print(f"  [OK] Ciclo criado com ID: {ciclo_id}")
    print(f"       Nome: {ciclo['nome']}")
    return ciclo_id


# ---------------------------------------------------------------------------
# 2. Ler e processar planilha
# ---------------------------------------------------------------------------

def processar_planilha(ciclo_id):
    """Lê o Excel e retorna lista de registros prontos para inserção."""
    print()
    print("=" * 60)
    print(f"PASSO 2: Lendo planilha: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    print(f"  Sheet: '{ws.title.strip()}'")
    print(f"  Total de linhas na sheet: {ws.max_row}")
    print()

    registros = []
    pulados = []
    total_lido = 0

    for row_num in range(4, ws.max_row + 1):
        # Lê células das colunas B a J (2 a 10)
        col_b = ws.cell(row=row_num, column=2).value   # operador
        col_c = ws.cell(row=row_num, column=3).value   # data
        col_d = ws.cell(row=row_num, column=4).value   # codigo_produto
        col_e = ws.cell(row=row_num, column=5).value   # lote
        col_f = ws.cell(row=row_num, column=6).value   # qtd_sistemica
        col_g = ws.cell(row=row_num, column=7).value   # qtd_fisica
        col_h = ws.cell(row=row_num, column=8).value   # qtd_venda
        # col_i = col 9 → ignorar (estoque — será recalculado)
        col_j = ws.cell(row=row_num, column=10).value  # observacao

        # Pular linhas completamente vazias
        if all(v is None for v in [col_b, col_c, col_d, col_e, col_f, col_g, col_h, col_j]):
            continue

        total_lido += 1

        # --- Validar e normalizar código produto ---
        codigo = normalizar_codigo(col_d)
        if not codigo:
            motivo = f"Linha {row_num}: código produto vazio"
            pulados.append(motivo)
            print(f"  [SKIP] {motivo}")
            continue

        # --- Validar e normalizar lote ---
        lote = normalizar_lote(col_e)
        if not lote:
            motivo = f"Linha {row_num}: lote vazio"
            pulados.append(motivo)
            print(f"  [SKIP] {motivo}")
            continue

        # --- Normalizar data ---
        data_iso, erro_data = normalizar_data(col_c, row_num)
        if data_iso is None:
            motivo = erro_data or f"Linha {row_num}: data inválida"
            pulados.append(motivo)
            print(f"  [SKIP] {motivo}")
            continue

        # --- Normalizar operador ---
        operador = normalizar_operador(col_b) or "Desconhecido"

        # --- Quantidades ---
        qtd_sistemica = to_int(col_f, default=0)
        qtd_fisica = to_int(col_g, default=None)  # pode ficar None
        qtd_venda = to_int(col_h, default=0)       # None → 0

        # --- Calcular divergência ---
        if qtd_fisica is not None:
            qtd_divergencia = qtd_fisica - (qtd_sistemica - qtd_venda)
        else:
            qtd_divergencia = None

        # --- Calcular percentual de divergência ---
        if qtd_divergencia is not None and qtd_sistemica and qtd_sistemica > 0:
            pct_divergencia = round(abs(qtd_divergencia) / qtd_sistemica * 100, 2)
        else:
            pct_divergencia = None

        # --- Status ---
        # Tabela aceita apenas "OK" e "AJUSTE_APROVADO"
        # Linhas sem qtd_fisica (divergencia None) = dados históricos sem contagem física → OK
        if qtd_divergencia is None or qtd_divergencia == 0:
            status = "OK"
        else:
            status = "AJUSTE_APROVADO"  # dados históricos já tratados

        # --- Observação ---
        observacao = str(col_j).strip() if col_j is not None else None
        if observacao == "None" or observacao == "":
            observacao = None

        registro = {
            "ciclo_id": ciclo_id,
            "codigo_produto": codigo,
            "lote": lote,
            "operador_nome": operador,
            "qtd_sistemica": qtd_sistemica,
            "qtd_fisica": qtd_fisica,
            "qtd_venda": qtd_venda,
            "qtd_divergencia": qtd_divergencia,
            "pct_divergencia": pct_divergencia,
            "status": status,
            "observacao": observacao,
            "contado_em": data_iso,
        }

        registros.append(registro)

    print(f"  Linhas lidas (não vazias): {total_lido}")
    print(f"  Registros válidos para inserção: {len(registros)}")
    print(f"  Registros pulados: {len(pulados)}")

    return registros, pulados, total_lido


# ---------------------------------------------------------------------------
# 3. Inserir em lotes
# ---------------------------------------------------------------------------

def inserir_em_lotes(registros):
    """Insere os registros na tabela inventario_contagens em lotes de BATCH_SIZE."""
    print()
    print("=" * 60)
    print(f"PASSO 3: Inserindo {len(registros)} registros em lotes de {BATCH_SIZE}...")

    total_inserido = 0
    erros = []

    for i in range(0, len(registros), BATCH_SIZE):
        lote = registros[i:i + BATCH_SIZE]
        lote_num = i // BATCH_SIZE + 1
        total_lotes = (len(registros) + BATCH_SIZE - 1) // BATCH_SIZE

        print(f"  Lote {lote_num}/{total_lotes} ({len(lote)} registros)...", end=" ")

        ok, resultado = supabase_post("inventario_contagens", lote)
        if ok:
            inseridos = len(resultado) if isinstance(resultado, list) else len(lote)
            total_inserido += inseridos
            print(f"OK ({inseridos} inseridos)")
        else:
            print(f"ERRO")
            erros.append(f"Lote {lote_num}: {resultado}")
            print(f"    Detalhe: {resultado[:200]}")

    return total_inserido, erros


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print()
    print("=" * 60)
    print("  IMPORTACAO HISTORICO INVENTARIO CONTINUO -> SUPABASE")
    print("=" * 60)
    print()

    # 1. Criar ciclo
    ciclo_id = criar_ciclo()

    # 2. Processar planilha
    registros, pulados, total_lido = processar_planilha(ciclo_id)

    if not registros:
        print()
        print("[AVISO] Nenhum registro válido para inserir. Encerrando.")
        sys.exit(0)

    # 3. Inserir em lotes
    total_inserido, erros = inserir_em_lotes(registros)

    # 4. Resumo final
    print()
    print("=" * 60)
    print("RESUMO FINAL")
    print("=" * 60)
    print(f"  Total de linhas lidas (não vazias) : {total_lido}")
    print(f"  Total de registros inseridos       : {total_inserido}")
    print(f"  Total de linhas puladas            : {len(pulados)}")

    if pulados:
        print()
        print("  Motivos dos registros pulados:")
        for motivo in pulados:
            print(f"    - {motivo}")

    if erros:
        print()
        print(f"  Erros de inserção ({len(erros)} lotes com falha):")
        for err in erros:
            print(f"    - {err}")

    print()
    if not erros:
        print("  [SUCESSO] Importação concluída sem erros!")
    else:
        print("  [ATENÇÃO] Importação concluída com erros. Verifique acima.")

    print()


if __name__ == "__main__":
    main()
