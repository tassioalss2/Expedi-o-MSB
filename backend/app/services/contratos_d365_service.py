"""Importa os contratos de venda exportados do D365.

Existe para responder "de qual cliente é esta solicitação?". O e-mail da
licitação quase sempre cita o contrato MSB ("MSB-000238"), e o contrato aponta o
código do cliente — medido, é a chave que mais rende: 81 dos 214 casos, à frente
do CNPJ lido no anexo (66) e da nota de empenho que já tem demanda (31).

O arquivo é o export padrão do D365 (`DynamicsExport_*.xlsx`) com as colunas
"ID do contrato de venda", "Conta de cliente", "Nome", "Título do documento" e
"Status". O motor reimporta o que estiver em `Licitacao/D365/` a cada rodada,
então contrato novo entra sozinho.
"""
import re
from datetime import datetime, timezone
from typing import Optional

from app.core.database import get_service_db

# "PE 90124/2025 DIVERSOS", "PE 63/2026 FIO GUIA", "PE 0404/26 FIO GUIA".
# O prefixo varia (PE, PP, pregão, dispensa) e às vezes não existe, então o que
# ancora é o formato numero/ano — não a sigla.
_PREGAO_NO_TITULO = re.compile(r'\b(\d{2,6}\s*/\s*\d{2,4})\b')

_COLUNAS = {
    "contrato": "ID do contrato de venda",
    "codigo_cliente": "Conta de cliente",
    "nome_d365": "Nome",
    "titulo": "Título do documento",
    "status": "Status",
}


def pregao_do_titulo(titulo: Optional[str]) -> Optional[str]:
    m = _PREGAO_NO_TITULO.search(str(titulo or ""))
    return re.sub(r"\s+", "", m.group(1)) if m else None


def e_export_de_contratos(caminho: str) -> bool:
    """Se o arquivo é um export de CONTRATOS, pelas colunas — não pelo nome.

    A pasta `Licitacao/D365/` recebe vários exports do D365 com nome parecido
    (`DynamicsExport_*.xlsx`): notas fiscais para a conciliação, contratos para
    cá. O nome não distingue, e a conciliação já resolve isso do mesmo jeito
    (`tipo_do_arquivo` olha o cabeçalho). Escolher pelo mais recente sem olhar
    dentro faria o importador tentar ler um export de notas e não encontrar
    contrato nenhum.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        cab = [str(c or "").strip() for c in next(ws.iter_rows(values_only=True), ())]
        wb.close()
    except Exception:
        return False
    return _COLUNAS["contrato"] in cab and _COLUNAS["codigo_cliente"] in cab


def importar(caminho: str) -> dict:
    """Lê o xlsx e grava. Idempotente: o contrato é a chave primária."""
    import openpyxl

    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    linhas = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    if not linhas:
        return {"lidos": 0, "gravados": 0, "sem_cliente": 0, "erro": "planilha vazia"}

    cab = [str(c or "").strip() for c in linhas[0]]
    faltando = [nome for nome in _COLUNAS.values() if nome not in cab]
    if faltando:
        # Recusa em silêncio é pior: se o D365 mudar o cabeçalho, alguém precisa
        # saber por que o cliente parou de ser resolvido.
        return {"lidos": 0, "gravados": 0, "sem_cliente": 0,
                "erro": "colunas ausentes: %s" % ", ".join(faltando)}
    idx = {chave: cab.index(nome) for chave, nome in _COLUNAS.items()}

    db = get_service_db()
    # Código do cliente → id. Paginado: o cadastro tem ~3.900 clientes e o
    # PostgREST devolve no máximo 1.000 por vez.
    por_codigo: dict = {}
    passo, ini = 1000, 0
    while True:
        bloco = db.table("clientes").select("id, codigo").limit(passo).offset(ini).execute().data
        for c in bloco:
            if c.get("codigo"):
                por_codigo[str(c["codigo"]).strip().upper()] = c["id"]
        ini += passo
        if len(bloco) < passo:
            break

    agora = datetime.now(timezone.utc).isoformat()
    registros, sem_cliente = [], 0
    for linha in linhas[1:]:
        if not linha:
            continue
        contrato = str(linha[idx["contrato"]] or "").strip()
        if not contrato:
            continue
        codigo = str(linha[idx["codigo_cliente"]] or "").strip()
        cliente_id = por_codigo.get(codigo.upper())
        if not cliente_id:
            sem_cliente += 1
        titulo = str(linha[idx["titulo"]] or "").strip() or None
        registros.append({
            "contrato": contrato.upper(),
            "codigo_cliente": codigo,
            "cliente_id": cliente_id,
            "nome_d365": str(linha[idx["nome_d365"]] or "").strip() or None,
            "titulo": titulo,
            "pregao": pregao_do_titulo(titulo),
            "status": str(linha[idx["status"]] or "").strip() or None,
            "importado_em": agora,
        })

    gravados = 0
    for i in range(0, len(registros), 100):
        lote = registros[i:i + 100]
        # Upsert em lote pela chave primária. O cliente deste projeto declara
        # `upsert(data: dict)`, mas manda o corpo como JSON e com
        # `resolution=merge-duplicates` — então lista funciona e o PostgREST
        # infere o alvo do conflito pela PK, sem precisar de `on_conflict`.
        db.table("licitacao_contratos_d365").upsert(lote).execute()
        gravados += len(lote)

    return {"lidos": len(registros), "gravados": gravados,
            "sem_cliente": sem_cliente, "erro": None}


def mapa_para_resolucao(db) -> tuple[dict, dict]:
    """(contrato → cliente_id, pregão → cliente_id).

    O pregão só entra quando aponta um cliente ÚNICO. Há pregões compartilhados
    por mais de um contrato — e quando eles são de clientes diferentes, escolher
    um seria atribuir a venda ao hospital errado.
    """
    rows = db.table("licitacao_contratos_d365")\
        .select("contrato, cliente_id, pregao").limit(5000).execute().data
    por_contrato = {r["contrato"]: r["cliente_id"] for r in rows if r.get("cliente_id")}
    candidatos: dict = {}
    for r in rows:
        if r.get("pregao") and r.get("cliente_id"):
            candidatos.setdefault(r["pregao"], set()).add(r["cliente_id"])
    por_pregao = {p: next(iter(c)) for p, c in candidatos.items() if len(c) == 1}
    return por_contrato, por_pregao
