"""Carga do histórico de faturamento item a item do D365 (planilhas de export).

Duas planilhas, dois papéis:

  faturamento_2025_2026.xlsx (aba Export) — o QUE foi vendido, para QUEM.
    Uma linha por item de fatura: cliente, tipo de cliente, UF, cidade,
    vertical, família, produto, quantidade e receita. O export já vem filtrado
    com "UNIDADE_DE_NEGOCIO não é INTERCOMPANY", então o transfer price fica
    fora na origem — não precisa (nem deve) filtrar Biomedical aqui de novo.

  historico_faturamento.xlsx (aba Sheet1) — QUANTO CUSTOU.
    Movimentações de estoque por OV. As linhas com Saída = 'Vendido' têm
    quantidade e "Valor de custo físico" negativos; o custo unitário é a razão
    dos módulos.

POR QUE CUSTO MÉDIO E NÃO CUSTO DA VENDA: as duas planilhas não compartilham
chave. A de faturamento identifica pelo número da FATURA, a de custo pelo número
da OV. Dá para casar por produto, não por transação. Então o custo é a média das
saídas daquele produto — suficiente para margem por cliente, produto, segmento e
região, que é o uso; insuficiente para auditar a margem de uma NF específica.

A planilha traz uma linha 'Total' e um rodapé com os filtros aplicados no fim —
linhas sem cliente ou sem produto são descartadas, senão o faturamento dobra.
"""
from collections import defaultdict
from datetime import datetime
from typing import Optional

from app.core.database import get_service_db

# Índices 0-based das colunas do export de faturamento (aba Export).
_C_ANOMES, _C_DATA, _C_FATURA = 4, 5, 2
_C_VERTICAL = 8
_C_TIPO_CLI, _C_COD_CLI, _C_CLIENTE = 13, 14, 15
_C_UF, _C_CIDADE = 16, 17
_C_FAMILIA, _C_COD_PROD, _C_PROD = 20, 21, 22
_C_RECEITA, _C_QTD, _C_ASP = 23, 24, 25

# Índices do export de custo (aba Sheet1).
_H_ITEM, _H_SAIDA, _H_QTD, _H_CUSTO = 0, 6, 7, 8

# O D365 escreve o tipo com acento e espaço; normalizado para caber em código.
_TIPO_CLIENTE = {
    "DISTRIBUIDOR": "DISTRIBUIDOR",
    "ÓRGÃO PÚBLICO": "ORGAO_PUBLICO",
    "ORGAO PUBLICO": "ORGAO_PUBLICO",
    "VENDA DIRETA": "VENDA_DIRETA",
}

_LOTE = 200


def _norm_codigo(v) -> str:
    return str(v or "").strip().upper()


def _competencia(anomes) -> Optional[str]:
    """202607 -> '2026-07'."""
    s = str(anomes or "").strip()
    if len(s) != 6 or not s.isdigit():
        return None
    return f"{s[:4]}-{s[4:]}"


def _data_iso(v) -> Optional[str]:
    if isinstance(v, datetime):
        return v.date().isoformat()
    return None


def custo_por_produto(caminho_hist: str) -> dict:
    """produto_codigo -> {custo_unitario, amostras, custo_min, custo_max}."""
    import openpyxl
    wb = openpyxl.load_workbook(caminho_hist, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    amostras: dict = defaultdict(list)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[_H_SAIDA] != "Vendido":
            continue
        qtd, custo = r[_H_QTD], r[_H_CUSTO]
        if not qtd or not custo:
            continue
        q, c = abs(float(qtd)), abs(float(custo))
        if q <= 0:
            continue
        cod = _norm_codigo(r[_H_ITEM])
        if cod:
            amostras[cod].append(c / q)
    wb.close()

    out = {}
    for cod, vals in amostras.items():
        out[cod] = {
            "produto_codigo": cod,
            "custo_unitario": round(sum(vals) / len(vals), 4),
            "amostras": len(vals),
            "custo_min": round(min(vals), 4),
            "custo_max": round(max(vals), 4),
        }
    return out


def ler_faturamento(caminho_fat: str, custos: dict) -> list:
    """Linhas prontas para gravar, já com custo aplicado."""
    import openpyxl
    wb = openpyxl.load_workbook(caminho_fat, read_only=True, data_only=True)
    ws = wb["Export"]
    linhas = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        cliente = r[_C_CLIENTE]
        cod_prod = _norm_codigo(r[_C_COD_PROD])
        # Descarta a linha 'Total' e o rodapé de filtros: só linha com cliente E
        # produto é fato de venda. Sem isso o total dobra.
        if not cliente or not cod_prod:
            continue
        comp = _competencia(r[_C_ANOMES])
        if not comp:
            continue

        qtd = float(r[_C_QTD] or 0)
        receita = round(float(r[_C_RECEITA] or 0), 2)
        cu = (custos.get(cod_prod) or {}).get("custo_unitario")
        tipo_bruto = str(r[_C_TIPO_CLI] or "").strip().upper()

        linhas.append({
            "competencia": comp,
            "data_faturamento": _data_iso(r[_C_DATA]),
            "numero_fatura": str(r[_C_FATURA] or "").strip() or None,
            "cliente_codigo": str(r[_C_COD_CLI] or "").strip() or None,
            "cliente_nome": str(cliente).strip(),
            "cliente_tipo": _TIPO_CLIENTE.get(tipo_bruto),
            "uf": str(r[_C_UF] or "").strip() or None,
            "cidade": str(r[_C_CIDADE] or "").strip() or None,
            "vertical": str(r[_C_VERTICAL] or "").strip() or None,
            "familia": str(r[_C_FAMILIA] or "").strip() or None,
            "produto_codigo": cod_prod,
            "produto_descricao": str(r[_C_PROD] or "").strip() or None,
            "qtd": qtd,
            "receita": receita,
            "preco_medio": round(float(r[_C_ASP]), 4) if r[_C_ASP] else None,
            "custo_unitario": cu,
            "custo_total": round(cu * qtd, 2) if (cu is not None and qtd) else None,
        })
    wb.close()
    return linhas


def importar(caminho_fat: str, caminho_hist: str, substituir: bool = True) -> dict:
    """Carrega as duas planilhas. `substituir` limpa as competências presentes no
    arquivo antes de gravar — reimportar o mesmo mês não duplica."""
    db = get_service_db()

    custos = custo_por_produto(caminho_hist)
    if custos:
        # Custo é sempre substituição integral: é a foto do custo médio atual.
        try:
            for cod in list(custos.keys()):
                db.table("produto_custo").delete().eq("produto_codigo", cod).execute()
        except Exception:
            pass
        vals = list(custos.values())
        for i in range(0, len(vals), _LOTE):
            db.table("produto_custo").insert(vals[i:i + _LOTE]).execute()

    linhas = ler_faturamento(caminho_fat, custos)
    competencias = sorted({l["competencia"] for l in linhas})

    if substituir:
        for comp in competencias:
            db.table("faturamento_itens").delete().eq("competencia", comp).execute()

    for i in range(0, len(linhas), _LOTE):
        db.table("faturamento_itens").insert(linhas[i:i + _LOTE]).execute()

    receita = round(sum(l["receita"] for l in linhas), 2)
    custo = round(sum(l["custo_total"] or 0 for l in linhas), 2)
    sem_custo = sorted({l["produto_codigo"] for l in linhas if l["custo_unitario"] is None})
    return {
        "linhas": len(linhas),
        "competencias": competencias,
        "produtos_com_custo": len(custos),
        "produtos_sem_custo": sem_custo,
        "receita": receita,
        "custo": custo,
        "margem_pct": round((receita - custo) / receita * 100, 1) if receita else None,
    }
